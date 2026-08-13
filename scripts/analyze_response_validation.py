"""Aggregate psychometric diagnostics for the 2025 Culture Fit responses.

The script intentionally exports no respondent-level rows or identifiers.
It also treats the scoring key below as a provisional content-audit key, not
as a validated production key.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter
from datetime import time
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from scipy.stats import chi2
from sklearn.decomposition import FactorAnalysis
from sklearn.feature_extraction.text import TfidfVectorizer


CORE_EXTRA_REVERSE = {
    "I039",
    "I047",
    "I071",
    "I079",
    "I101",
    "I102",
    "I110",
    "I269",
    "I270",
    "I271",
    "I272",
    "I276",
    "I282",
}
CORE_FALSE_REVERSE_SUFFIX = {"I055_R", "I118_R"}
INTEGRITY_OVERT_REVERSE = {
    "I181",
    "I182",
    "I183",
    "I184",
    "I185",
    "I186",
    "I189",
    "I190",
}


def normalized_text(value: str) -> str:
    return re.sub(r"[\W_]+", "", str(value).lower(), flags=re.UNICODE)


def numeric_item_order(item_id: str) -> int:
    digits = re.sub(r"\D", "", str(item_id))
    return int(digits) if digits else 0


def provisional_reverse(item: pd.Series) -> bool:
    item_id = str(item["item_id"])
    variable = str(item["variable"])
    if variable == "Culture-Fit":
        if item_id in CORE_FALSE_REVERSE_SUFFIX:
            return False
        return item_id.endswith("_R") or item_id in CORE_EXTRA_REVERSE
    if variable == "CWB":
        return True
    if variable == "Integrity (overt)":
        return item_id in INTEGRITY_OVERT_REVERSE
    if variable == "Social Desirability (rev)":
        return True
    return False


def current_app_reverse(item: pd.Series) -> bool:
    variable = str(item["variable"])
    domain = str(item["domain"])
    return (
        variable in {"CWB", "Integrity (overt)", "Social Desirability (rev)"}
        or "역기능행동" in domain
        or "CWB" in domain
    )


def safe_correlation(left: pd.Series, right: pd.Series) -> float:
    if left.std(ddof=0) == 0 or right.std(ddof=0) == 0:
        return float("nan")
    return float(left.corr(right))


def cronbach_alpha(frame: pd.DataFrame) -> float:
    clean = frame.dropna(axis=0)
    k = clean.shape[1]
    if k < 2 or clean.shape[0] < 3:
        return float("nan")
    item_variances = clean.var(axis=0, ddof=1).sum()
    total_variance = clean.sum(axis=1).var(ddof=1)
    if total_variance <= 0:
        return float("nan")
    return float(k / (k - 1) * (1 - item_variances / total_variance))


def mean_interitem_correlation(frame: pd.DataFrame) -> float:
    corr = frame.corr().to_numpy()
    values = corr[np.triu_indices_from(corr, 1)]
    values = values[np.isfinite(values)]
    return float(values.mean()) if len(values) else float("nan")


def omega_one_factor(frame: pd.DataFrame) -> float:
    clean = frame.dropna(axis=0)
    if clean.shape[0] < 10 or clean.shape[1] < 3:
        return float("nan")
    z = (clean - clean.mean()) / clean.std(ddof=0).replace(0, np.nan)
    z = z.dropna(axis=1)
    if z.shape[1] < 3:
        return float("nan")
    model = FactorAnalysis(n_components=1, random_state=2026).fit(z)
    loadings = model.components_[0].copy()
    if loadings.sum() < 0:
        loadings *= -1
    numerator = float(loadings.sum() ** 2)
    denominator = numerator + float(model.noise_variance_.sum())
    return numerator / denominator if denominator > 0 else float("nan")


def kmo_and_bartlett(frame: pd.DataFrame) -> tuple[float, float, float]:
    clean = frame.dropna(axis=0)
    clean = clean.loc[:, clean.std(ddof=0) > 0]
    n, p = clean.shape
    if n <= p + 1 or p < 3:
        return float("nan"), float("nan"), float("nan")
    corr = clean.corr().to_numpy()
    if not np.all(np.isfinite(corr)):
        return float("nan"), float("nan"), float("nan")
    inv = np.linalg.pinv(corr)
    diag = np.sqrt(np.clip(np.diag(inv), 1e-12, None))
    partial = -inv / np.outer(diag, diag)
    np.fill_diagonal(partial, 0)
    corr_off = corr.copy()
    np.fill_diagonal(corr_off, 0)
    r2 = float(np.sum(corr_off**2))
    p2 = float(np.sum(partial**2))
    kmo = r2 / (r2 + p2) if r2 + p2 else float("nan")

    sign, logdet = np.linalg.slogdet(corr)
    if sign <= 0:
        return kmo, float("nan"), float("nan")
    statistic = -(n - 1 - (2 * p + 5) / 6) * logdet
    dof = p * (p - 1) / 2
    return kmo, float(statistic), float(chi2.sf(statistic, dof))


def parallel_factor_count(frame: pd.DataFrame, repetitions: int = 200) -> int:
    clean = frame.dropna(axis=0)
    clean = clean.loc[:, clean.std(ddof=0) > 0]
    n, p = clean.shape
    if n < 20 or p < 3:
        return 0
    observed = np.linalg.eigvalsh(clean.corr().to_numpy())[::-1]
    rng = np.random.default_rng(20260730)
    random_eigenvalues = np.empty((repetitions, p))
    for i in range(repetitions):
        random_eigenvalues[i] = np.linalg.eigvalsh(
            np.corrcoef(rng.normal(size=(n, p)), rowvar=False)
        )[::-1]
    threshold = np.percentile(random_eigenvalues, 95, axis=0)
    return int(np.sum(observed > threshold))


def item_rest_correlations(frame: pd.DataFrame) -> pd.Series:
    result = {}
    total = frame.sum(axis=1)
    for column in frame:
        rest = total - frame[column]
        result[column] = safe_correlation(frame[column], rest)
    return pd.Series(result, dtype=float)


def seconds_from_excel(value: object) -> float:
    if isinstance(value, time):
        return value.hour * 3600 + value.minute * 60 + value.second
    if isinstance(value, pd.Timedelta):
        return value.total_seconds()
    if isinstance(value, str):
        parts = value.split(":")
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + float(parts[2])
    return float("nan")


def imc_correct(item_id: str, answer: int, timestamp: pd.Timestamp) -> bool:
    exact = {
        "I251": 1,
        "I252": 3,
        "I253": 5,
        "I256": 1,
        "I257": 1,
        "I258": 1,
        "I259": 3,
        "I260": 5,
    }
    if item_id == "I254":
        return answer != 5
    if item_id == "I255":
        if pd.isna(timestamp):
            return answer == 5
        seoul = timestamp.tz_convert("Asia/Seoul") if timestamp.tzinfo else timestamp
        return seoul.weekday() == 0 or answer == 5
    return answer == exact[item_id]


def read_jobfit_questions(path: Path) -> pd.DataFrame:
    workbook = load_workbook(path, read_only=True, data_only=False)
    sheet = workbook.worksheets[-1]
    records = []
    for row in sheet.iter_rows(min_row=13, values_only=True):
        if row[0] and str(row[0]).startswith("Q"):
            records.append(
                {
                    "question_id": row[0],
                    "job_group": row[1],
                    "area": row[2],
                    "element": row[3],
                    "text": row[4],
                }
            )
    return pd.DataFrame(records)


def top_text_matches(
    jobfit: pd.DataFrame, items: pd.DataFrame
) -> tuple[pd.DataFrame, pd.DataFrame]:
    corpus = jobfit["text"].fillna("").tolist() + items["text"].fillna("").tolist()
    vectorizer = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5), min_df=1)
    matrix = vectorizer.fit_transform(corpus)
    job_matrix = matrix[: len(jobfit)]
    culture_matrix = matrix[len(jobfit) :]
    similarities = (job_matrix @ culture_matrix.T).toarray()
    rows = []
    for i, question in jobfit.iterrows():
        for j in np.argsort(similarities[i])[-3:][::-1]:
            rows.append(
                {
                    "question_id": question["question_id"],
                    "job_group": question["job_group"],
                    "area": question["area"],
                    "element": question["element"],
                    "culture_item_id": items.iloc[j]["item_id"],
                    "culture_domain": items.iloc[j]["domain"],
                    "similarity": similarities[i, j],
                }
            )

    within = (job_matrix @ job_matrix.T).toarray()
    pair_rows = []
    for i in range(len(jobfit)):
        for j in range(i + 1, len(jobfit)):
            pair_rows.append(
                {
                    "question_id_1": jobfit.iloc[i]["question_id"],
                    "job_group_1": jobfit.iloc[i]["job_group"],
                    "question_id_2": jobfit.iloc[j]["question_id"],
                    "job_group_2": jobfit.iloc[j]["job_group"],
                    "similarity": within[i, j],
                }
            )
    pairs = pd.DataFrame(pair_rows).sort_values("similarity", ascending=False)
    return pd.DataFrame(rows), pairs


def rounded_record(record: dict) -> dict:
    clean = {}
    for key, value in record.items():
        if isinstance(value, (np.floating, float)):
            clean[key] = None if not math.isfinite(float(value)) else round(float(value), 4)
        elif isinstance(value, (np.integer,)):
            clean[key] = int(value)
        else:
            clean[key] = value
    return clean


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--responses", default="Culture Fit_Responses.xlsx")
    parser.add_argument("--items", default="data/items_full.json")
    parser.add_argument(
        "--jobfit", default="job ontology_2026(AI 작성) - 복사본.xlsx"
    )
    parser.add_argument("--output", default="analysis_outputs")
    args = parser.parse_args()

    output = Path(args.output)
    output.mkdir(parents=True, exist_ok=True)

    items = pd.DataFrame(
        json.loads(Path(args.items).read_text(encoding="utf-8"))
    ).reset_index(drop=True)
    # The 2025 submit2.js used a stable numeric sort of item_id before writing
    # answers. Reproducing that rule is essential because N001 sorts beside I001.
    storage_order = sorted(
        items.index, key=lambda index: numeric_item_order(items.loc[index, "item_id"])
    )
    question_by_index = {
        item_index: f"Q{position}"
        for position, item_index in enumerate(storage_order, start=1)
    }
    items["question"] = items.index.map(question_by_index)
    items["reverse_provisional"] = items.apply(provisional_reverse, axis=1)
    items["reverse_current_app"] = items.apply(current_app_reverse, axis=1)
    items["normalized_text"] = items["text"].map(normalized_text)

    responses = pd.read_excel(args.responses, sheet_name="Responses")
    question_columns = items["question"].tolist()
    raw = responses[question_columns].apply(pd.to_numeric, errors="coerce")
    scored = raw.copy()
    for _, item in items[items["reverse_provisional"]].iterrows():
        scored[item["question"]] = 6 - scored[item["question"]]
    current_scored = raw.copy()
    for _, item in items[items["reverse_current_app"]].iterrows():
        current_scored[item["question"]] = 6 - current_scored[item["question"]]

    duplicate_rows = []
    duplicate_question_sets = []
    for group_id, (_, group) in enumerate(
        items.groupby("normalized_text", sort=False), start=1
    ):
        if len(group) < 2:
            continue
        questions = group["question"].tolist()
        duplicate_question_sets.append(questions)
        pair_corr = []
        pair_agreement = []
        pair_abs_diff = []
        for i in range(len(questions)):
            for j in range(i + 1, len(questions)):
                a, b = raw[questions[i]], raw[questions[j]]
                pair_corr.append(safe_correlation(a, b))
                pair_agreement.append((a == b).mean())
                pair_abs_diff.append((a - b).abs().mean())
        duplicate_rows.append(
            {
                "group_id": group_id,
                "item_ids": ", ".join(group["item_id"]),
                "domain": ", ".join(sorted(group["domain"].unique())),
                "copies": len(group),
                "mean_pair_correlation": np.nanmean(pair_corr),
                "mean_exact_agreement": np.mean(pair_agreement),
                "mean_absolute_difference": np.mean(pair_abs_diff),
                "text": group.iloc[0]["text"],
            }
        )
    duplicate_stats = pd.DataFrame(duplicate_rows)

    unique_items = items.drop_duplicates("normalized_text", keep="first")
    scale_rows = []
    item_rows = []
    scale_names = items["domain"].drop_duplicates().tolist()
    for scale in scale_names:
        scale_items = items[items["domain"] == scale]
        unique_scale_items = unique_items[unique_items["domain"] == scale]
        frame = scored[scale_items["question"].tolist()]
        unique_frame = scored[unique_scale_items["question"].tolist()]
        current_frame = current_scored[scale_items["question"].tolist()]
        current_unique_frame = current_scored[
            unique_scale_items["question"].tolist()
        ]
        rest = item_rest_correlations(frame)
        kmo, bartlett, bartlett_p = kmo_and_bartlett(unique_frame)
        scale_rows.append(
            {
                "scale": scale,
                "n": len(frame),
                "items_full": frame.shape[1],
                "items_unique": unique_frame.shape[1],
                "alpha_current_app_full": cronbach_alpha(current_frame),
                "alpha_current_app_unique": cronbach_alpha(current_unique_frame),
                "alpha_full": cronbach_alpha(frame),
                "alpha_unique": cronbach_alpha(unique_frame),
                "omega_unique_one_factor": omega_one_factor(unique_frame),
                "mean_interitem_r_unique": mean_interitem_correlation(unique_frame),
                "kmo_unique": kmo,
                "bartlett_chi2_unique": bartlett,
                "bartlett_p_unique": bartlett_p,
                "parallel_factors_unique": parallel_factor_count(unique_frame),
            }
        )
        for _, item in scale_items.iterrows():
            raw_values = raw[item["question"]]
            values = scored[item["question"]]
            item_rows.append(
                {
                    "question": item["question"],
                    "item_id": item["item_id"],
                    "domain": item["domain"],
                    "variable": item["variable"],
                    "scoring_key_provisional": (
                        "reverse" if item["reverse_provisional"] else "direct"
                    ),
                    "mean_raw": raw_values.mean(),
                    "mean_scored": values.mean(),
                    "sd_scored": values.std(ddof=1),
                    "pct_raw_1": (raw_values == 1).mean(),
                    "pct_raw_5": (raw_values == 5).mean(),
                    "item_rest_r": rest[item["question"]],
                    "exact_duplicate": bool(
                        (items["normalized_text"] == item["normalized_text"]).sum() > 1
                    ),
                    "text": item["text"],
                }
            )

    scale_stats = pd.DataFrame(scale_rows)
    item_stats = pd.DataFrame(item_rows)
    item_stats["review_flags"] = item_stats.apply(
        lambda row: ";".join(
            flag
            for flag, condition in [
                ("LOW_ITEM_REST", row["item_rest_r"] < 0.20),
                ("NEGATIVE_ITEM_REST", row["item_rest_r"] < 0),
                ("LOW_VARIANCE", row["sd_scored"] < 0.50),
                ("RAW_CEILING", row["pct_raw_5"] >= 0.60),
                ("RAW_FLOOR", row["pct_raw_1"] >= 0.60),
                ("EXACT_DUPLICATE", row["exact_duplicate"]),
            ]
            if condition
        ),
        axis=1,
    )

    core_domains = [
        "원칙중시",
        "혁신성",
        "고객중심",
        "의사소통",
        "도전정신",
    ]
    domain_scores = {}
    scoring_sensitivity_rows = []
    for domain in core_domains:
        cols = unique_items[
            (unique_items["domain"] == domain)
            & (unique_items["variable"] == "Culture-Fit")
        ]["question"].tolist()
        domain_scores[domain] = scored[cols].mean(axis=1)
        current_domain_score = current_scored[cols].mean(axis=1)
        provisional_domain_score = scored[cols].mean(axis=1)
        scoring_sensitivity_rows.append(
            {
                "domain": domain,
                "items_unique": len(cols),
                "current_app_mean": current_domain_score.mean(),
                "provisional_mean": provisional_domain_score.mean(),
                "mean_absolute_person_difference": (
                    current_domain_score - provisional_domain_score
                ).abs().mean(),
                "pearson_r": safe_correlation(
                    current_domain_score, provisional_domain_score
                ),
                "spearman_rank_r": current_domain_score.corr(
                    provisional_domain_score, method="spearman"
                ),
            }
        )
    domain_correlations = pd.DataFrame(domain_scores).corr()
    scoring_sensitivity = pd.DataFrame(scoring_sensitivity_rows)

    timestamps = pd.to_datetime(responses["Timestamp"], errors="coerce", utc=True)
    imc_items = items[items["is_imc"]]
    imc_failures = []
    imc_item_rows = []
    imc_pass_matrix = {}
    for _, item in imc_items.iterrows():
        passes = []
        for row_index in range(len(responses)):
            answer = int(raw.loc[row_index, item["question"]])
            passes.append(
                imc_correct(item["item_id"], answer, timestamps.iloc[row_index])
            )
        imc_pass_matrix[item["item_id"]] = passes
        imc_item_rows.append(
            {
                "item_id": item["item_id"],
                "question": item["question"],
                "pass_rate": np.mean(passes),
                "failures": int(len(passes) - sum(passes)),
                "text": item["text"],
            }
        )
    for row_index in range(len(responses)):
        failures = sum(
            not imc_pass_matrix[item_id][row_index]
            for item_id in imc_pass_matrix
        )
        imc_failures.append(failures)
    imc_item_stats = pd.DataFrame(imc_item_rows)

    duration = responses["Time Spent"].map(seconds_from_excel)
    row_max_share = raw.apply(
        lambda row: row.value_counts(normalize=True).max(), axis=1
    )
    person_sd = raw.std(axis=1, ddof=1)
    duplicate_agreement_per_person = pd.Series(0.0, index=raw.index)
    duplicate_comparisons = 0
    for questions in duplicate_question_sets:
        for i in range(len(questions)):
            for j in range(i + 1, len(questions)):
                duplicate_agreement_per_person += (
                    raw[questions[i]] == raw[questions[j]]
                ).astype(float)
                duplicate_comparisons += 1
    duplicate_agreement_per_person /= duplicate_comparisons

    quality = {
        "rows": len(responses),
        "completed_rows": int((responses["Status"] == "COMPLETED").sum()),
        "complete_300_rows": int((responses["Completion"] == "300/300").sum()),
        "missing_answer_cells": int(raw.isna().sum().sum()),
        "invalid_answer_cells": int(
            ((~raw.isin([1, 2, 3, 4, 5])) & raw.notna()).sum().sum()
        ),
        "missing_session_ids": int(responses["Session ID"].isna().sum()),
        "forced_submits": int((responses["Forced Submit"] != "NO").sum()),
        "duration_seconds_median": duration.median(),
        "duration_seconds_p10": duration.quantile(0.10),
        "duration_seconds_p90": duration.quantile(0.90),
        "seconds_per_item_median": duration.median() / len(items),
        "under_10_minutes": int((duration < 600).sum()),
        "under_1_5_seconds_per_item": int((duration / len(items) < 1.5).sum()),
        "focus_out_median": pd.to_numeric(
            responses["Focus Out Count"], errors="coerce"
        ).median(),
        "focus_out_over_10": int(
            (
                pd.to_numeric(responses["Focus Out Count"], errors="coerce")
                > 10
            ).sum()
        ),
        "row_max_response_share_median": row_max_share.median(),
        "row_max_response_share_over_80pct": int((row_max_share > 0.8).sum()),
        "person_sd_below_0_5": int((person_sd < 0.5).sum()),
        "imc_failure_mean": np.mean(imc_failures),
        "imc_failure_distribution": dict(
            sorted(Counter(imc_failures).items())
        ),
        "imc_any_failure": int((np.array(imc_failures) > 0).sum()),
        "imc_two_or_more_failures": int((np.array(imc_failures) >= 2).sum()),
        "duplicate_agreement_person_median": duplicate_agreement_per_person.median(),
        "duplicate_agreement_person_below_80pct": int(
            (duplicate_agreement_per_person < 0.8).sum()
        ),
    }

    jobfit = read_jobfit_questions(Path(args.jobfit))
    jobfit_matches, jobfit_pairs = top_text_matches(jobfit, items)
    jobfit_summary = {
        "questions": len(jobfit),
        "job_groups": int(jobfit["job_group"].nunique()),
        "job_group_counts": jobfit["job_group"].value_counts().to_dict(),
        "area_counts": jobfit["area"].value_counts().to_dict(),
        "all_positive_keyed": True,
        "top_match_similarity_ge_0_50": int(
            (
                jobfit_matches.groupby("question_id")["similarity"].max()
                >= 0.50
            ).sum()
        ),
        "within_jobfit_pairs_ge_0_60": int(
            (jobfit_pairs["similarity"] >= 0.60).sum()
        ),
    }

    duplicate_stats.to_csv(
        output / "exact_duplicate_response_diagnostics.csv",
        index=False,
        encoding="utf-8-sig",
    )
    scale_stats.to_csv(
        output / "scale_psychometric_summary.csv",
        index=False,
        encoding="utf-8-sig",
    )
    item_stats.to_csv(
        output / "item_statistics.csv", index=False, encoding="utf-8-sig"
    )
    domain_correlations.to_csv(
        output / "core_domain_correlations.csv", encoding="utf-8-sig"
    )
    scoring_sensitivity.to_csv(
        output / "scoring_key_sensitivity.csv",
        index=False,
        encoding="utf-8-sig",
    )
    imc_item_stats.to_csv(
        output / "imc_item_diagnostics.csv",
        index=False,
        encoding="utf-8-sig",
    )
    jobfit.to_csv(
        output / "jobfit_question_inventory.csv",
        index=False,
        encoding="utf-8-sig",
    )
    jobfit_matches.to_csv(
        output / "jobfit_culturefit_similarity.csv",
        index=False,
        encoding="utf-8-sig",
    )
    jobfit_pairs.head(100).to_csv(
        output / "jobfit_internal_similarity_top100.csv",
        index=False,
        encoding="utf-8-sig",
    )

    summary = {
        "scoring_key_status": "provisional_content_audit_key",
        "response_quality": rounded_record(quality),
        "duplicate_groups": len(duplicate_stats),
        "duplicate_surplus_items": int(
            sum(len(group) - 1 for group in duplicate_question_sets)
        ),
        "duplicate_pair_correlation_median": float(
            duplicate_stats["mean_pair_correlation"].median()
        ),
        "duplicate_pair_agreement_median": float(
            duplicate_stats["mean_exact_agreement"].median()
        ),
        "item_review_flag_counts": Counter(
            flag
            for flags in item_stats["review_flags"]
            for flag in flags.split(";")
            if flag
        ),
        "jobfit": jobfit_summary,
    }
    summary = json.loads(
        json.dumps(summary, ensure_ascii=False, default=float)
    )
    (output / "analysis_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
